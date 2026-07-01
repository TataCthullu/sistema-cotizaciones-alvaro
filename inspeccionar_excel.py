"""from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Monto"
ws["E12"] = 1000

wb.save("prueba.xlsx")

print("Excel creado")"""

import openpyxl

ARCHIVO = r"d:/escritos/codigo/Arbitraje_web/backend/exl_alv.xlsx"

wb = openpyxl.load_workbook(ARCHIVO, data_only=False)

for nombre_hoja in wb.sheetnames:
    hoja = wb[nombre_hoja]
    print(f"\n===== HOJA: {nombre_hoja} =====")
    print(f"Dimensiones: {hoja.dimensions}")
    print(f"Filas con datos: {hoja.max_row} | Columnas: {hoja.max_column}")

    # Mostrar encabezados (primera fila) y hasta 15 filas de datos
    max_filas_mostrar = min(hoja.max_row, 16)  # fila 1 + 15 de datos
    for fila in hoja.iter_rows(min_row=1, max_row=max_filas_mostrar, values_only=False):
        celdas_info = []
        for celda in fila:
            if celda.value is None:
                celdas_info.append("")
            elif isinstance(celda.value, str) and celda.value.startswith("="):
                celdas_info.append(f"[F] {celda.value}")  # F de fórmula
            else:
                celdas_info.append(str(celda.value))
        print(" | ".join(celdas_info))

    # Si hay listas desplegables (validación de datos)
    if hoja.data_validations and hoja.data_validations.dataValidation:
        print("\n Validaciones de datos (listas desplegables):")
        for dv in hoja.data_validations.dataValidation:
            print(f"   Rango: {dv.sqref}, Tipo: {dv.type}, Fórmula1: {dv.formula1}")

    # Si hay celdas combinadas
    if hoja.merged_cells.ranges:
        print("\n Celdas combinadas:")
        for rango in hoja.merged_cells.ranges:
            print(f"   {rango}")

    # Anchura de columnas (a veces indica estructura)
    print("\n Ancho de columnas:")
    for columna, dim in hoja.column_dimensions.items():
        if dim.width:
            print(f"   Col {columna}: ancho {dim.width}")

print("\n===== FIN DEL REPORTE =====")